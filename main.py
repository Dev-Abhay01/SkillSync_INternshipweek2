import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time

# Global Constants
ATTENDANCE_FILE = 'attendance.xlsx'
ATTENDANCE_THRESHOLD = 0.75
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
SENDER_EMAIL = "sender_email@gmail.com"
SENDER_PASSWORD = "password"
EMAIL_LIST = []

def save_excel(workbook):
    """
    Save the updated Excel sheet.
    """
    workbook.save(ATTENDANCE_FILE)

def track_attendance(worksheet):
    """
    Track attendance and send an email alert if attendance falls below the threshold.
    """
    total_students = worksheet.max_row
    absent_students = worksheet.max_column

    attendance_ratio = (total_students - absent_students) / total_students

    if attendance_ratio < ATTENDANCE_THRESHOLD:
        print(f"Attendance ratio is {attendance_ratio * 100}%. Sending an email alert.")
        for email in EMAIL_LIST:
            send_email(email)
    else:
        print(f"Attendance ratio is {attendance_ratio * 100}%. No action required.")

def send_email(to_email):
    """
    Send an email alert.
    """
    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
        server.starttls()
        server.login(SENDER_EMAIL, SENDER_PASSWORD)

        msg = MIMEMultipart()
        msg['From'] = SENDER_EMAIL
        msg['To'] = to_email
        msg['Subject'] = "Attendance Alert"

        body = "Dear Admin,\n\nThe attendance ratio is below the threshold.\n\nBest Regards,\nThe Attendance Tracker"
        msg.attach(MIMEText(body, 'plain'))

        server.sendmail(SENDER_EMAIL, to_email, msg.as_string())

        print(f"Email sent to {to_email}")

if __name__ == "__main__":
    workbook = openpyxl.load_workbook(ATTENDANCE_FILE)
    worksheet = workbook.active

    while True:
        track_attendance(worksheet)
        save_excel(workbook)
        time.sleep(3600)  # 1 hour
